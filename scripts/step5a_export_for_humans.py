#!/usr/bin/env python3
"""
TOC-Bench Step 5a (v2): Export for Human Verification
======================================================
Takes the combined filter output (_combined.json) from step4 v2 and
produces:

  1. toc_bench_clean.json     — final benchmark JSON (polished text,
                                 internal debug fields stripped)
  2. annotation_sheet.xlsx    — Excel for human annotators, 4 sheets
                                 (MCQ / Ordering / Statement Pair /
                                  Numerical) + Summary
  3. consolidate_videos.sh    — shell script to copy/convert all videos
                                 into a flat folder as .mp4

Key v2 contracts:
  - Reads polished_* fields written by step3c v2 (polished_question,
    polished_option_A..D, polished_statement_A/B, polished_events[i]
    .event_text). Falls back to raw fields if polished not present.
  - Uses v2 schema fields: dim, tier, subject_label (singular),
    hallucination, has_hallucination_distractor. Drops v1 fields
    (task_type, phenomenon, dimension, subject_labels).
  - Supports all 4 v2 formats: mcq_4, sp, ordering_3/4, numerical.
  - Hallucination items are exported normally (masked) so annotators
    treat them like ordinary questions. The hallucination tag is
    preserved in the clean JSON for post-hoc agreement analysis.

Usage:
    python step5a_export_for_humans.py
    python step5a_export_for_humans.py --input path/to/_combined.json
    python step5a_export_for_humans.py --sample 1000  # random stratified subset
"""

import argparse
import json
import random
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from configs.config import QA_DIR, TRACKS_DIR, DIMENSIONS

FILTER_DIR = QA_DIR / "filtered_natural"
EXPORT_DIR = QA_DIR / "human_verification_natural"


# ============================================================
# Polished-field readers (v2)
# ============================================================

def _pq(item):
    return (item.get("polished_question")
            or item.get("question") or "").strip()


def _po(item, letter):
    return (item.get(f"polished_option_{letter}")
            or item.get(f"option_{letter}") or "").strip()


def _ps(item, suffix):
    return (item.get(f"polished_statement_{suffix}")
            or item.get(f"statement_{suffix}") or "").strip()


def _pe(item):
    """Return ordering events list with polished event_text where available."""
    polished = item.get("polished_events")
    if polished:
        return [
            {"label": e["label"],
             "event_text": (e.get("event_text") or "").strip()}
            for e in polished
        ]
    return [
        {"label": e["label"],
         "event_text": (e.get("event_text") or "").strip()}
        for e in item.get("events", [])
    ]


# ============================================================
# Clean Benchmark JSON (public-facing, no internal debug fields)
# ============================================================

# Internal fields we deliberately strip from the public benchmark JSON
# (they're useful for step3c/step4 debugging but not for evaluation).
_STRIP_FIELDS = {
    "realization_method", "verification_passed", "verification_issues",
    "layer1_result", "layer2_result", "layer3_result",
    "skeleton_id",
    # step3c/b internal distractor-tracking
    "distractor_type_tag",
    # step3a unit provenance
    "unit_id",
}


def clean_item(item):
    """
    Strip internal/debug fields, keep only benchmark-facing fields.
    Uses polished_* text when available. Preserves v2 metadata
    (dim, tier, hallucination) under `metadata`.
    """
    fmt = item.get("format", "mcq_4")

    base = {
        "qa_id": item["qa_id"],
        "video_id": item["video_id"],
        "format": fmt,
        "question": _pq(item),
    }

    if fmt == "mcq_4":
        base["option_A"] = _po(item, "A")
        base["option_B"] = _po(item, "B")
        base["option_C"] = _po(item, "C")
        base["option_D"] = _po(item, "D")
        base["correct_answer"] = item.get("correct_answer", "")

    elif fmt == "sp":
        base["statement_A"] = _ps(item, "A")
        base["statement_B"] = _ps(item, "B")
        base["correct_answer"] = item.get("correct_answer", "")

    elif fmt.startswith("ordering_"):
        base["events"] = _pe(item)
        base["correct_order"] = item.get("correct_order", [])

    elif fmt == "numerical":
        # Open-ended numerical. correct_answer is a string in
        # {"2","3","4","5 or more"}.
        base["correct_answer"] = item.get("correct_answer", "")
        base["allowed_answers"] = ["2", "3", "4", "5 or more"]

    # v2 metadata (kept for analysis, not for evaluation).
    # Hallucination tag is preserved so we can check per-variant
    # human agreement, but it is not surfaced in the Excel sheet
    # given to annotators (see masking policy in docstring).
    base["metadata"] = {
        "dim": item.get("dim"),
        "tier": item.get("tier"),
        "subject_label": item.get("subject_label"),
        "hallucination": item.get("hallucination"),  # None / variant_a / variant_b
        "has_hallucination_distractor": bool(
            item.get("has_hallucination_distractor", False)
        ),
    }

    return base


# ============================================================
# Excel Export (4 format-specific sheets + Summary)
# ============================================================

def build_excel(clean_items, output_path, subset_label=""):
    """Create an annotation Excel workbook with one sheet per format."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---------- Style helpers ----------
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    answer_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
    issue_fill = PatternFill("solid", fgColor="FCE4EC")    # light pink
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, n_cols):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

    def style_row(ws, row, n_cols, answer_cols=(), issue_col=None):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.alignment = wrap
            c.border = border
            if col in answer_cols:
                c.fill = answer_fill
            if issue_col and col == issue_col:
                c.fill = issue_fill

    # Partition items by format
    mcq_items = [it for it in clean_items if it["format"] == "mcq_4"]
    sp_items = [it for it in clean_items if it["format"] == "sp"]
    ord_items = [it for it in clean_items if it["format"].startswith("ordering_")]
    num_items = [it for it in clean_items if it["format"] == "numerical"]

    def _meta_cols(it):
        """Return (dim, tier) for display in an Excel row. Kept short so
        the sheet stays readable."""
        md = it.get("metadata") or {}
        return md.get("dim", ""), md.get("tier", "")

    # ---------- Sheet: MCQ ----------
    ws = wb.active
    ws.title = "MCQ"
    headers = ["No.", "QA ID", "Video ID", "Dim", "Tier", "Question",
               "A", "B", "C", "D",
               "Your Answer\n(A/B/C/D)", "Issue\n(1=问题)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    for i, it in enumerate(mcq_items):
        r = i + 2
        dim, tier = _meta_cols(it)
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=it["qa_id"])
        ws.cell(row=r, column=3, value=it["video_id"])
        ws.cell(row=r, column=4, value=dim)
        ws.cell(row=r, column=5, value=tier)
        ws.cell(row=r, column=6, value=it["question"])
        ws.cell(row=r, column=7, value=it.get("option_A", ""))
        ws.cell(row=r, column=8, value=it.get("option_B", ""))
        ws.cell(row=r, column=9, value=it.get("option_C", ""))
        ws.cell(row=r, column=10, value=it.get("option_D", ""))
        style_row(ws, r, len(headers), answer_cols={11}, issue_col=12)

    for col, w in enumerate(
        [6, 14, 20, 22, 8, 45, 26, 26, 26, 26, 15, 12], 1
    ):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # ---------- Sheet: Statement Pair ----------
    ws = wb.create_sheet("Statement Pair")
    headers = ["No.", "QA ID", "Video ID", "Dim", "Tier", "Question",
               "Statement A", "Statement B",
               "Your Answer\n(A/B)", "Issue\n(1=问题)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    for i, it in enumerate(sp_items):
        r = i + 2
        dim, tier = _meta_cols(it)
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=it["qa_id"])
        ws.cell(row=r, column=3, value=it["video_id"])
        ws.cell(row=r, column=4, value=dim)
        ws.cell(row=r, column=5, value=tier)
        ws.cell(row=r, column=6, value=it["question"])
        ws.cell(row=r, column=7, value=it.get("statement_A", ""))
        ws.cell(row=r, column=8, value=it.get("statement_B", ""))
        style_row(ws, r, len(headers), answer_cols={9}, issue_col=10)

    for col, w in enumerate(
        [6, 14, 20, 22, 8, 45, 38, 38, 15, 12], 1
    ):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # ---------- Sheet: Ordering ----------
    ws = wb.create_sheet("Ordering")
    # ordering_3 has events A/B/C; ordering_4 has A/B/C/D.
    # We show 4 event columns; D is blank for ordering_3.
    headers = ["No.", "QA ID", "Video ID", "Dim", "Tier", "Question",
               "Event A", "Event B", "Event C", "Event D",
               "Your Order\n(如 C,A,B 或 C,A,B,D)", "Issue\n(1=问题)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    for i, it in enumerate(ord_items):
        r = i + 2
        dim, tier = _meta_cols(it)
        events = {e["label"]: e["event_text"] for e in it.get("events", [])}
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=it["qa_id"])
        ws.cell(row=r, column=3, value=it["video_id"])
        ws.cell(row=r, column=4, value=dim)
        ws.cell(row=r, column=5, value=tier)
        ws.cell(row=r, column=6, value=it["question"])
        ws.cell(row=r, column=7, value=events.get("A", ""))
        ws.cell(row=r, column=8, value=events.get("B", ""))
        ws.cell(row=r, column=9, value=events.get("C", ""))
        ws.cell(row=r, column=10, value=events.get("D", ""))
        style_row(ws, r, len(headers), answer_cols={11}, issue_col=12)

    for col, w in enumerate(
        [6, 14, 20, 22, 8, 38, 28, 28, 28, 28, 20, 12], 1
    ):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # ---------- Sheet: Numerical ----------
    ws = wb.create_sheet("Numerical")
    headers = ["No.", "QA ID", "Video ID", "Dim", "Tier", "Question",
               "Your Count\n(2 / 3 / 4 / 5 or more)", "Issue\n(1=问题)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    for i, it in enumerate(num_items):
        r = i + 2
        dim, tier = _meta_cols(it)
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=it["qa_id"])
        ws.cell(row=r, column=3, value=it["video_id"])
        ws.cell(row=r, column=4, value=dim)
        ws.cell(row=r, column=5, value=tier)
        ws.cell(row=r, column=6, value=it["question"])
        style_row(ws, r, len(headers), answer_cols={7}, issue_col=8)

    for col, w in enumerate([6, 14, 20, 22, 8, 50, 22, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # ---------- Sheet: Summary (inserted at position 0) ----------
    ws = wb.create_sheet("Summary", 0)
    title = "TOC-Bench Human Verification (v2)"
    if subset_label:
        title = f"TOC-Bench Human Verification — {subset_label} (v2)"
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=14)

    ws["A3"] = "Instructions:"
    ws["A3"].font = Font(bold=True)
    instructions = [
        "1. Watch the video (filename = <Video ID>.mp4 in the videos folder).",
        "2. Read the question and options carefully.",
        "3. Fill in your answer in the YELLOW column only:",
        "   - MCQ sheet: enter A, B, C, or D",
        "   - Statement Pair: enter A or B",
        "   - Ordering: enter the chronological order, e.g. C,A,B (ordering_3) "
        "or C,A,B,D (ordering_4)",
        "   - Numerical: enter 2, 3, 4, or exactly \"5 or more\"",
        "4. If there is ANY issue with the question (unclear, wrong options, no "
        "correct answer, bad grammar, etc.), put 1 in the PINK Issue column.",
        "5. Do NOT modify any other columns.",
        "",
        "Note: some items have no physically valid answer in the video "
        "(e.g. they ask about an object that is not present). For those, "
        "choose the option that best reflects what you actually see — do "
        "not infer from context.",
    ]
    for i, line in enumerate(instructions):
        ws.cell(row=4 + i, column=1, value=line)

    cur_row = 4 + len(instructions) + 2

    # Totals
    ws.cell(row=cur_row, column=1, value="Statistics:").font = Font(bold=True)
    cur_row += 1
    ws.cell(row=cur_row, column=1,
            value=f"Total QA items: {len(clean_items):,}")
    cur_row += 1
    ws.cell(row=cur_row, column=1, value=f"  MCQ: {len(mcq_items):,}")
    cur_row += 1
    ws.cell(row=cur_row, column=1,
            value=f"  Statement Pair: {len(sp_items):,}")
    cur_row += 1
    ws.cell(row=cur_row, column=1, value=f"  Ordering: {len(ord_items):,}")
    cur_row += 1
    ws.cell(row=cur_row, column=1, value=f"  Numerical: {len(num_items):,}")
    cur_row += 1
    vid_count = len(set(it["video_id"] for it in clean_items))
    ws.cell(row=cur_row, column=1, value=f"  Unique videos: {vid_count:,}")
    cur_row += 2

    # Per-tier distribution
    ws.cell(row=cur_row, column=1, value="By tier:").font = Font(bold=True)
    cur_row += 1
    tier_counts = Counter(
        (it.get("metadata") or {}).get("tier", "?") for it in clean_items
    )
    for tier in ["tier1", "tier2", "tier3"]:
        cnt = tier_counts.get(tier, 0)
        pct = cnt / max(1, len(clean_items)) * 100
        ws.cell(row=cur_row, column=1,
                value=f"  {tier}: {cnt:,} ({pct:.1f}%)")
        cur_row += 1
    cur_row += 1

    # Per-dim distribution
    ws.cell(row=cur_row, column=1, value="By dim:").font = Font(bold=True)
    cur_row += 1
    dim_counts = Counter(
        (it.get("metadata") or {}).get("dim", "?") for it in clean_items
    )
    for dim, cnt in dim_counts.most_common():
        cfg = DIMENSIONS.get(dim, {})
        c_mark = "★" if cfg.get("c_critical") else " "
        pct = cnt / max(1, len(clean_items)) * 100
        ws.cell(row=cur_row, column=1,
                value=f"  {c_mark} {dim}: {cnt:,} ({pct:.1f}%)")
        cur_row += 1

    ws.column_dimensions["A"].width = 80

    wb.save(output_path)
    print(f"  Excel saved: {output_path}")
    print(f"    MCQ sheet:            {len(mcq_items):>6,d} rows")
    print(f"    Statement Pair sheet: {len(sp_items):>6,d} rows")
    print(f"    Ordering sheet:       {len(ord_items):>6,d} rows")
    print(f"    Numerical sheet:      {len(num_items):>6,d} rows")


# ============================================================
# Video Consolidation Script
# ============================================================

def build_video_script(clean_items, output_path, video_dest_dir):
    """Generate a shell script that copies/converts all needed videos
    into a single flat folder as .mp4 files."""
    video_ids = sorted(set(it["video_id"] for it in clean_items))

    lines = [
        "#!/bin/bash",
        "# TOC-Bench: Consolidate videos for human verification",
        f"# {len(video_ids)} unique videos to process",
        "",
        f'DEST="{video_dest_dir}"',
        'mkdir -p "$DEST"',
        "",
        "DONE=0",
        "FAIL=0",
        "",
    ]

    for vid in video_ids:
        track_path = TRACKS_DIR / f"{vid}.json"
        if not track_path.exists():
            lines.append(f'echo "[SKIP] {vid}: no track file"')
            lines.append("FAIL=$((FAIL+1))")
            continue

        try:
            with open(track_path) as f:
                td = json.load(f)
            video_path = td.get("video_path", "")
        except Exception:
            lines.append(f'echo "[SKIP] {vid}: track file unreadable"')
            lines.append("FAIL=$((FAIL+1))")
            continue

        if not video_path:
            lines.append(f'echo "[SKIP] {vid}: no video_path in track"')
            lines.append("FAIL=$((FAIL+1))")
            continue

        p = Path(video_path)
        dest_mp4 = f'"$DEST"/{vid}.mp4'

        if p.suffix == "" or p.is_dir():
            lines.append(f'# {vid}: image folder → mp4')
            lines.append(f'if [ -d "{video_path}" ]; then')
            lines.append(
                f'  ffmpeg -y -framerate 12 -pattern_type glob '
                f'-i "{video_path}/*.jpg" '
                f'-c:v libx264 -pix_fmt yuv420p -q:v 2 '
                f'{dest_mp4} -loglevel warning 2>/dev/null'
            )
            lines.append("  if [ $? -ne 0 ]; then")
            lines.append(
                f'    ffmpeg -y -framerate 12 -pattern_type glob '
                f'-i "{video_path}/*.png" '
                f'-c:v libx264 -pix_fmt yuv420p -q:v 2 '
                f'{dest_mp4} -loglevel warning 2>/dev/null'
            )
            lines.append("  fi")
            lines.append(
                f'  if [ -f {dest_mp4} ]; then DONE=$((DONE+1)); '
                f'else echo "[FAIL] {vid}"; FAIL=$((FAIL+1)); fi'
            )
            lines.append("else")
            lines.append(
                f'  echo "[SKIP] {vid}: dir not found"; FAIL=$((FAIL+1))'
            )
            lines.append("fi")
        elif p.suffix.lower() == ".mp4":
            lines.append(f'# {vid}: copy mp4')
            lines.append(f'if [ -f "{video_path}" ]; then')
            lines.append(
                f'  cp "{video_path}" {dest_mp4} && DONE=$((DONE+1))'
            )
            lines.append("else")
            lines.append(
                f'  echo "[SKIP] {vid}: file not found"; FAIL=$((FAIL+1))'
            )
            lines.append("fi")
        else:
            lines.append(f'# {vid}: {p.suffix} → mp4')
            lines.append(f'if [ -f "{video_path}" ]; then')
            lines.append(
                f'  ffmpeg -y -i "{video_path}" '
                f'-c:v libx264 -pix_fmt yuv420p -q:v 2 '
                f'{dest_mp4} -loglevel warning && DONE=$((DONE+1))'
            )
            lines.append("else")
            lines.append(
                f'  echo "[SKIP] {vid}: file not found"; FAIL=$((FAIL+1))'
            )
            lines.append("fi")

        lines.append("")

    lines.extend([
        'echo ""',
        'echo "==============================="',
        f'echo "  Done: $DONE / {len(video_ids)}"',
        'echo "  Failed: $FAIL"',
        'echo "==============================="',
    ])

    with open(output_path, "w") as f:
        f.write("\n".join(lines) + "\n")

    import os
    os.chmod(output_path, 0o755)

    print(f"  Video script saved: {output_path}")
    print(f"    {len(video_ids)} unique videos → {video_dest_dir}")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="TOC-Bench Step 5a v2: Export for human verification"
    )
    parser.add_argument("--input", type=str, default=None,
                        help="Path to _combined.json (default: "
                             f"{FILTER_DIR}/_combined.json)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help=f"Destination directory for outputs (default: "
                             f"{EXPORT_DIR}). Pass a subset-specific path "
                             f"when exporting multiple subsets "
                             f"(e.g. verified vs quality_sample) to avoid "
                             f"overwriting each other.")
    parser.add_argument("--subset-label", type=str, default="",
                        help="Short label for this subset (e.g. 'verified' "
                             "or 'quality_sample'). Used in the Excel "
                             "Summary title and the clean JSON version "
                             "string so downstream tools can distinguish "
                             "the two exports.")
    parser.add_argument("--sample", type=int, default=None,
                        help="Random stratified sample size")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument(
        "--video-dir", type=str,
        default=str(QA_DIR.parent / "videos_for_annotation"),
        help="Destination directory for consolidated videos",
    )
    args = parser.parse_args()

    print("=" * 70)
    print("  TOC-Bench Step 5a v2: Export for Human Verification")
    if args.subset_label:
        print(f"  Subset: {args.subset_label}")
    print("=" * 70)

    # Resolve output directory (CLI override wins; default is package-level
    # EXPORT_DIR). Both verified and quality_sample exports should pass
    # different --output-dir values so files don't clobber each other.
    output_dir = Path(args.output_dir) if args.output_dir else EXPORT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"  Output: {output_dir}")

    input_path = (Path(args.input) if args.input
                  else FILTER_DIR / "_combined.json")
    if not input_path.exists():
        print(f"  [ERROR] {input_path} not found. Run step4 --combine first.")
        sys.exit(1)

    with open(input_path) as f:
        data = json.load(f)
    items = data.get("qa_items", [])
    print(f"  Loaded {len(items):,} passed QA items from {input_path.name}")

    # Filter out dims that are marked deprecated in config. These dims are
    # preserved in the config for historical / negative-result reasons but
    # should not appear in the final benchmark.
    deprecated_dims = {
        name for name, cfg in DIMENSIONS.items() if cfg.get("deprecated")
    }
    if deprecated_dims:
        before = len(items)
        items = [it for it in items
                 if it.get("dim") not in deprecated_dims]
        n_removed = before - len(items)
        if n_removed > 0:
            print(f"  Excluded {n_removed:,} items from deprecated dims: "
                  f"{sorted(deprecated_dims)}")
        else:
            print(f"  Deprecated dims {sorted(deprecated_dims)} — 0 items "
                  f"to exclude (already absent)")

    # Stratified sampling by format
    if args.sample and args.sample < len(items):
        rng = random.Random(args.seed)
        by_fmt = {}
        for it in items:
            by_fmt.setdefault(it.get("format", "mcq_4"), []).append(it)

        sampled = []
        for fmt, group in by_fmt.items():
            n = max(1, round(args.sample * len(group) / len(items)))
            n = min(n, len(group))
            sampled.extend(rng.sample(group, n))
        rng.shuffle(sampled)
        items = sampled[:args.sample]
        print(f"  Stratified sample: {len(items):,} items")

    # Format breakdown
    fmt_counts = Counter(it.get("format", "?") for it in items)
    print(f"  Format breakdown:")
    for fmt, cnt in fmt_counts.most_common():
        print(f"    {fmt:<14s} {cnt:>6,d}")
    vid_count = len(set(it["video_id"] for it in items))
    print(f"  Unique videos: {vid_count:,}")

    # Build outputs
    clean_items = [clean_item(it) for it in items]

    # Version string includes subset label so downstream tooling and
    # annotators can tell exported subsets apart.
    version_str = "v2-pre-human-verification"
    if args.subset_label:
        version_str = f"{version_str}-{args.subset_label}"

    # 1. Clean benchmark JSON
    clean_path = output_dir / "toc_bench_clean.json"
    with open(clean_path, "w") as f:
        json.dump({
            "benchmark": "TOC-Bench",
            "version": version_str,
            "subset": args.subset_label or None,
            "total_items": len(clean_items),
            "format_counts": dict(fmt_counts),
            "video_count": vid_count,
            "deprecated_dims_excluded": sorted(deprecated_dims) if deprecated_dims else [],
            "items": clean_items,
        }, f, indent=2)
    print(f"\n  Clean JSON saved: {clean_path}")

    # 2. Annotation Excel
    excel_path = output_dir / "annotation_sheet.xlsx"
    build_excel(clean_items, excel_path, subset_label=args.subset_label)

    # 3. Video consolidation script
    script_path = output_dir / "consolidate_videos.sh"
    build_video_script(clean_items, script_path, args.video_dir)

    # Summary
    print(f"\n{'='*70}")
    print(f"  Export Complete")
    print(f"{'='*70}")
    print(f"  Clean JSON:       {clean_path}")
    print(f"  Annotation Excel: {excel_path}")
    print(f"  Video Script:     {script_path}")
    print(f"\n  Next steps:")
    print(f"    1. Run: bash {script_path}")
    print(f"    2. Distribute annotation_sheet.xlsx to annotators")
    print(f"    3. Collect completed sheets and run step5c to merge")


if __name__ == "__main__":
    main()